import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.contrib.admin.views.decorators import staff_member_required
from django.utils.decorators import method_decorator
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from orders.models import Order
from decimal import Decimal
from datetime import datetime

BLUE_FILL = PatternFill("solid", fgColor="2563EB")
ORANGE_FILL = PatternFill("solid", fgColor="F97316")
GREEN_FILL = PatternFill("solid", fgColor="16A34A")
LIGHT_FILL = PatternFill("solid", fgColor="F3F4F6")
WHITE_FONT = Font(color="FFFFFF", bold=True, size=11)
BOLD_FONT = Font(bold=True, size=10)
NORMAL_FONT = Font(size=10)

def style_header(cell, fill=None):
    cell.font = WHITE_FONT
    cell.fill = fill or BLUE_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

def style_cell(cell, bold=False, align="left"):
    cell.font = BOLD_FONT if bold else NORMAL_FONT
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

def add_univerin_header(ws, title, subtitle=""):
    ws.merge_cells("A1:H1")
    ws["A1"] = "UNIVERIN PRIVATE LIMITED"
    ws["A1"].font = Font(bold=True, size=16, color="2563EB")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:H2")
    ws["A2"] = "4/11, Sankarapuram, Govindampalli, Obulavaripalle - 516105, AP | GSTIN: 37AADCU8846J1ZP"
    ws["A2"].font = Font(size=9, color="6B7280")
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:H3")
    ws["A3"] = title
    ws["A3"].font = Font(bold=True, size=13)
    ws["A3"].alignment = Alignment(horizontal="center")

    if subtitle:
        ws.merge_cells("A4:H4")
        ws["A4"] = subtitle
        ws["A4"].font = Font(size=10, color="6B7280")
        ws["A4"].alignment = Alignment(horizontal="center")

    return 6  # Start data from row 6


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def seller_monthly_excel(request):
    """Seller downloads their monthly earnings Excel"""
    try:
        vendor = request.user.vendor
    except Exception:
        from rest_framework.response import Response
        return Response({"error": "Vendor not found"}, status=404)

    month = int(request.GET.get("month", datetime.now().month))
    year = int(request.GET.get("year", datetime.now().year))

    orders = Order.objects.filter(
        vendor=vendor, status="delivered",
        created_at__month=month, created_at__year=year
    ).order_by("created_at")

    wb = openpyxl.Workbook()

    # ── Sheet 1: Order Summary ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Order Summary"
    month_name = datetime(year, month, 1).strftime("%B %Y")
    start_row = add_univerin_header(ws1, f"Monthly Earnings Statement — {month_name}", f"Seller: {vendor.shop_name}")

    headers = ["Order ID", "Date", "Order Value", "Commission", "TCS (1%)", "Total Deductions", "Net Earnings", "Status"]
    col_widths = [20, 15, 15, 15, 12, 18, 15, 12]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=start_row, column=col, value=h)
        style_header(cell)
        ws1.column_dimensions[get_column_letter(col)].width = w

    ws1.row_dimensions[start_row].height = 30

    total_ov = Decimal("0"); total_comm = Decimal("0")
    total_tcs = Decimal("0"); total_net = Decimal("0")

    for i, order in enumerate(orders):
        row = start_row + 1 + i
        ov   = Decimal(str(order.subtotal or 0))
        comm = Decimal(str(order.commission_amount or 0))
        tcs  = Decimal(str(order.tcs_amount or 0))
        ded  = comm + tcs
        net  = ov - ded
        total_ov += ov; total_comm += comm
        total_tcs += tcs; total_net += net

        fill = LIGHT_FILL if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        row_data = [
            str(order.id)[:12].upper(),
            order.created_at.strftime("%d %b %Y"),
            f"Rs.{ov:.2f}",
            f"Rs.{comm:.2f}",
            f"Rs.{tcs:.2f}",
            f"Rs.{ded:.2f}",
            f"Rs.{net:.2f}",
            order.status.title()
        ]
        aligns = ["left","center","right","right","right","right","right","center"]
        for col, (val, align) in enumerate(zip(row_data, aligns), 1):
            cell = ws1.cell(row=row, column=col, value=val)
            style_cell(cell, align=align)
            cell.fill = fill

    # Total row
    total_row = start_row + 1 + len(orders)
    total_data = ["TOTAL", "", f"Rs.{total_ov:.2f}", f"Rs.{total_comm:.2f}", f"Rs.{total_tcs:.2f}", f"Rs.{total_comm+total_tcs:.2f}", f"Rs.{total_net:.2f}", ""]
    for col, val in enumerate(total_data, 1):
        cell = ws1.cell(row=total_row, column=col, value=val)
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = GREEN_FILL
        cell.alignment = Alignment(horizontal="right" if col > 2 else "center", vertical="center")
        cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    # Summary box
    summary_row = total_row + 2
    ws1.cell(row=summary_row, column=1, value="SUMMARY").font = Font(bold=True, size=12)
    summary_data = [
        ("Gross Order Value", f"Rs.{total_ov:.2f}"),
        ("Total Commission", f"Rs.{total_comm:.2f}"),
        ("Total TCS Deducted", f"Rs.{total_tcs:.2f}"),
        ("Total Deductions", f"Rs.{total_comm+total_tcs:.2f}"),
        ("NET PAYOUT", f"Rs.{total_net:.2f}"),
    ]
    for j, (label, value) in enumerate(summary_data):
        r = summary_row + 1 + j
        lc = ws1.cell(row=r, column=1, value=label)
        vc = ws1.cell(row=r, column=2, value=value)
        is_total = label == "NET PAYOUT"
        lc.font = Font(bold=is_total, size=10, color="16A34A" if is_total else "111827")
        vc.font = Font(bold=is_total, size=10, color="16A34A" if is_total else "111827")
        if is_total:
            lc.fill = PatternFill("solid", fgColor="F0FDF4")
            vc.fill = PatternFill("solid", fgColor="F0FDF4")

    # ── Sheet 2: GST Details ────────────────────────────────────────────
    ws2 = wb.create_sheet("GST Details")
    start_row2 = add_univerin_header(ws2, f"GST Details — {month_name}", f"Seller: {vendor.shop_name} | GSTIN: {getattr(vendor, 'gstin', 'N/A') or 'N/A'}")

    gst_headers = ["Order ID", "Date", "Taxable Value", "Commission", "GST on Commission (18%)", "CGST (9%)", "SGST (9%)", "TCS (1%)"]
    gst_widths  = [20, 15, 15, 15, 22, 12, 12, 12]
    for col, (h, w) in enumerate(zip(gst_headers, gst_widths), 1):
        cell = ws2.cell(row=start_row2, column=col, value=h)
        style_header(cell, ORANGE_FILL)
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.row_dimensions[start_row2].height = 30

    for i, order in enumerate(orders):
        row = start_row2 + 1 + i
        ov   = Decimal(str(order.subtotal or 0))
        comm = Decimal(str(order.commission_amount or 0))
        gst  = Decimal(str(order.gst_on_platform or 0))
        tcs  = Decimal(str(order.tcs_amount or 0))
        fill = LIGHT_FILL if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        row_data = [
            str(order.id)[:12].upper(),
            order.created_at.strftime("%d %b %Y"),
            f"Rs.{ov:.2f}",
            f"Rs.{comm:.2f}",
            f"Rs.{gst:.2f}",
            f"Rs.{gst/2:.2f}",
            f"Rs.{gst/2:.2f}",
            f"Rs.{tcs:.2f}",
        ]
        aligns = ["left","center","right","right","right","right","right","right"]
        for col, (val, align) in enumerate(zip(row_data, aligns), 1):
            cell = ws2.cell(row=row, column=col, value=val)
            style_cell(cell, align=align)
            cell.fill = fill

    # Save
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="univerin_earnings_{vendor.shop_name}_{month_name}.xlsx"'.replace(" ", "_")
    wb.save(response)
    return response


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def admin_billing_excel(request):
    """Admin downloads all billing data Excel"""
    # Support both admin session and JWT
    if not request.user.is_authenticated:
        from rest_framework.response import Response
        return Response({"error": "Login required"}, status=401)

    month = int(request.GET.get("month", datetime.now().month))
    year  = int(request.GET.get("year", datetime.now().year))
    month_name = datetime(year, month, 1).strftime("%B %Y")

    orders = Order.objects.filter(
        status="delivered",
        created_at__month=month,
        created_at__year=year
    ).select_related("vendor", "buyer").order_by("created_at")

    wb = openpyxl.Workbook()

    # ── Sheet 1: GSTR-1 Sales Register ─────────────────────────────────
    ws1 = wb.active
    ws1.title = "GSTR-1 Sales"
    start_row = add_univerin_header(ws1, f"GSTR-1 Sales Register — {month_name}", "Outward Supplies")
    headers = ["Order ID", "Date", "Buyer", "Vendor", "Taxable Value", "CGST", "SGST", "Total Invoice"]
    widths  = [20, 15, 20, 20, 15, 12, 12, 15]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws1.cell(row=start_row, column=col, value=h)
        style_header(cell)
        ws1.column_dimensions[get_column_letter(col)].width = w
    ws1.row_dimensions[start_row].height = 30

    t_taxable = Decimal("0"); t_total = Decimal("0")
    for i, order in enumerate(orders):
        row = start_row + 1 + i
        ov   = Decimal(str(order.subtotal or 0))
        gst  = Decimal(str(order.gst_on_platform or 0))
        total = ov + gst + Decimal(str(order.platform_fee or 0)) + Decimal(str(order.delivery_fee or 0))
        t_taxable += ov; t_total += total
        fill = LIGHT_FILL if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        row_data = [
            str(order.id)[:12].upper(),
            order.created_at.strftime("%d %b %Y"),
            getattr(order.buyer, "phone", "N/A"),
            order.vendor.shop_name,
            f"Rs.{ov:.2f}",
            f"Rs.{gst/2:.2f}",
            f"Rs.{gst/2:.2f}",
            f"Rs.{total:.2f}",
        ]
        aligns = ["left","center","left","left","right","right","right","right"]
        for col, (val, align) in enumerate(zip(row_data, aligns), 1):
            cell = ws1.cell(row=row, column=col, value=val)
            style_cell(cell, align=align)
            cell.fill = fill

    # GSTR-1 total
    tr = start_row + 1 + len(orders)
    for col, val in enumerate(["TOTAL", "", "", "", f"Rs.{t_taxable:.2f}", "", "", f"Rs.{t_total:.2f}"], 1):
        cell = ws1.cell(row=tr, column=col, value=val)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = BLUE_FILL
        cell.alignment = Alignment(horizontal="right" if col > 4 else "center")

    # ── Sheet 2: GSTR-8 TCS Register ───────────────────────────────────
    ws2 = wb.create_sheet("GSTR-8 TCS")
    start_row2 = add_univerin_header(ws2, f"GSTR-8 TCS Register — {month_name}", "Tax Collected at Source u/s 52 CGST Act")
    tcs_headers = ["Order ID", "Date", "Vendor", "Vendor GSTIN", "Taxable Value", "CGST TCS (0.5%)", "SGST TCS (0.5%)", "Total TCS"]
    tcs_widths  = [20, 15, 20, 20, 15, 18, 18, 12]
    for col, (h, w) in enumerate(zip(tcs_headers, tcs_widths), 1):
        cell = ws2.cell(row=start_row2, column=col, value=h)
        style_header(cell, ORANGE_FILL)
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.row_dimensions[start_row2].height = 30

    t_tcs = Decimal("0")
    for i, order in enumerate(orders):
        row = start_row2 + 1 + i
        ov  = Decimal(str(order.subtotal or 0))
        tcs = Decimal(str(order.tcs_amount or 0))
        t_tcs += tcs
        fill = LIGHT_FILL if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        row_data = [
            str(order.id)[:12].upper(),
            order.created_at.strftime("%d %b %Y"),
            order.vendor.shop_name,
            getattr(order.vendor, "gstin", "N/A") or "N/A",
            f"Rs.{ov:.2f}",
            f"Rs.{tcs/2:.2f}",
            f"Rs.{tcs/2:.2f}",
            f"Rs.{tcs:.2f}",
        ]
        aligns = ["left","center","left","left","right","right","right","right"]
        for col, (val, align) in enumerate(zip(row_data, aligns), 1):
            cell = ws2.cell(row=row, column=col, value=val)
            style_cell(cell, align=align)
            cell.fill = fill

    tr2 = start_row2 + 1 + len(orders)
    for col, val in enumerate(["TOTAL", "", "", "", "", "", "", f"Rs.{t_tcs:.2f}"], 1):
        cell = ws2.cell(row=tr2, column=col, value=val)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = ORANGE_FILL
        cell.alignment = Alignment(horizontal="right" if col > 4 else "center")

    # ── Sheet 3: Commission Register ────────────────────────────────────
    ws3 = wb.create_sheet("Commission Register")
    start_row3 = add_univerin_header(ws3, f"Commission Register — {month_name}", "Platform Commission & GST")
    comm_headers = ["Order ID", "Date", "Vendor", "Category", "Order Value", "Commission Rate", "Commission", "GST (18%)", "Total Commission"]
    comm_widths  = [20, 15, 20, 15, 15, 16, 14, 12, 18]
    for col, (h, w) in enumerate(zip(comm_headers, comm_widths), 1):
        cell = ws3.cell(row=start_row3, column=col, value=h)
        style_header(cell, GREEN_FILL)
        ws3.column_dimensions[get_column_letter(col)].width = w
    ws3.row_dimensions[start_row3].height = 30

    t_comm = Decimal("0"); t_gst_comm = Decimal("0")
    for i, order in enumerate(orders):
        row = start_row3 + 1 + i
        ov   = Decimal(str(order.subtotal or 0))
        comm = Decimal(str(order.commission_amount or 0))
        gst  = Decimal(str(order.gst_on_platform or 0))
        t_comm += comm; t_gst_comm += gst
        fill = LIGHT_FILL if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        row_data = [
            str(order.id)[:12].upper(),
            order.created_at.strftime("%d %b %Y"),
            order.vendor.shop_name,
            getattr(order.vendor, "category", "N/A") or "N/A",
            f"Rs.{ov:.2f}",
            f"{order.commission_rate or 0}%",
            f"Rs.{comm:.2f}",
            f"Rs.{gst:.2f}",
            f"Rs.{comm+gst:.2f}",
        ]
        aligns = ["left","center","left","left","right","center","right","right","right"]
        for col, (val, align) in enumerate(zip(row_data, aligns), 1):
            cell = ws3.cell(row=row, column=col, value=val)
            style_cell(cell, align=align)
            cell.fill = fill

    tr3 = start_row3 + 1 + len(orders)
    for col, val in enumerate(["TOTAL", "", "", "", "", "", f"Rs.{t_comm:.2f}", f"Rs.{t_gst_comm:.2f}", f"Rs.{t_comm+t_gst_comm:.2f}"], 1):
        cell = ws3.cell(row=tr3, column=col, value=val)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = GREEN_FILL
        cell.alignment = Alignment(horizontal="right" if col > 5 else "center")

    # ── Sheet 4: Settlement Summary ─────────────────────────────────────
    ws4 = wb.create_sheet("Settlement Summary")
    start_row4 = add_univerin_header(ws4, f"Settlement Summary — {month_name}", "Vendor wise Net Payout")
    sett_headers = ["Vendor", "Category", "Total Orders", "Gross Value", "Commission", "TCS", "Net Payout", "Bank A/C"]
    sett_widths  = [25, 15, 14, 15, 15, 12, 15, 20]
    for col, (h, w) in enumerate(zip(sett_headers, sett_widths), 1):
        cell = ws4.cell(row=start_row4, column=col, value=h)
        style_header(cell, PatternFill("solid", fgColor="7C3AED"))
        ws4.column_dimensions[get_column_letter(col)].width = w
    ws4.row_dimensions[start_row4].height = 30

    from django.db.models import Sum, Count
    vendor_summary = orders.values(
        "vendor__shop_name", "vendor__category", "vendor__shop_name"
    ).annotate(
        total_orders=Count("id"),
        gross=Sum("subtotal"),
        commission=Sum("commission_amount"),
        tcs=Sum("tcs_amount")
    )

    for i, v in enumerate(vendor_summary):
        row = start_row4 + 1 + i
        gross = Decimal(str(v["gross"] or 0))
        comm  = Decimal(str(v["commission"] or 0))
        tcs   = Decimal(str(v["tcs"] or 0))
        net   = gross - comm - tcs
        fill  = LIGHT_FILL if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        row_data = [
            v["vendor__shop_name"],
            v["vendor__category"] or "N/A",
            str(v["total_orders"]),
            f"Rs.{gross:.2f}",
            f"Rs.{comm:.2f}",
            f"Rs.{tcs:.2f}",
            f"Rs.{net:.2f}",
            v["vendor__shop_name"] or "N/A",
        ]
        aligns = ["left","left","center","right","right","right","right","left"]
        for col, (val, align) in enumerate(zip(row_data, aligns), 1):
            cell = ws4.cell(row=row, column=col, value=val)
            style_cell(cell, align=align)
            cell.fill = fill
            if col == 7:
                cell.font = Font(bold=True, size=10, color="16A34A")

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="univerin_admin_billing_{month_name}.xlsx"'.replace(" ", "_")
    wb.save(response)
    return response
