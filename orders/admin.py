from django.contrib import admin
from django.http import HttpResponse
from django.utils.html import format_html
from django.urls import path
from .models import Order, OrderItem, Notification, Review, Cart
from datetime import datetime

@admin.register(Order)
class OrderAdmin(admin.ModelAdmin):
    list_display  = ['order_number', 'buyer', 'vendor', 'status', 'subtotal', 'commission_amount', 'tcs_amount', 'created_at', 'download_invoice']
    list_filter   = ['status', 'payment_mode', 'created_at']
    search_fields = ['order_number', 'buyer__full_name', 'vendor__shop_name']
    ordering      = ['-created_at']

    def download_invoice(self, obj):
        return format_html(
            '<a href="/api/invoices/buyer/{}/" target="_blank" style="background:#2563eb;color:white;padding:4px 10px;border-radius:4px;text-decoration:none;font-size:12px;">📄 Invoice</a>',
            obj.id
        )
    download_invoice.short_description = "Invoice"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path("export-excel/", self.admin_site.admin_view(self.export_excel_view), name="order-export-excel"),
        ]
        return custom_urls + urls

    def export_excel_view(self, request):
        from invoices.excel_views import admin_billing_excel
        from django.http import HttpResponse
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from orders.models import Order
        from decimal import Decimal
        from datetime import datetime
        from django.db.models import Sum, Count

        month = int(request.GET.get("month", datetime.now().month))
        year  = int(request.GET.get("year", datetime.now().year))
        month_name = datetime(year, month, 1).strftime("%B %Y")

        orders = Order.objects.filter(
            status="delivered",
            created_at__month=month,
            created_at__year=year
        ).select_related("vendor", "buyer").order_by("created_at")

        wb = openpyxl.Workbook()

        BLUE_FILL   = PatternFill("solid", fgColor="2563EB")
        ORANGE_FILL = PatternFill("solid", fgColor="F97316")
        GREEN_FILL  = PatternFill("solid", fgColor="16A34A")
        LIGHT_FILL  = PatternFill("solid", fgColor="F3F4F6")
        WHITE_FONT  = Font(color="FFFFFF", bold=True, size=11)
        BOLD_FONT   = Font(bold=True, size=10)
        NORMAL_FONT = Font(size=10)

        def sh(cell, fill=None):
            cell.font = WHITE_FONT
            cell.fill = fill or BLUE_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

        def sc(cell, bold=False, align="left"):
            cell.font = BOLD_FONT if bold else NORMAL_FONT
            cell.alignment = Alignment(horizontal=align, vertical="center")
            cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

        # Sheet 1 GSTR-1
        ws1 = wb.active
        ws1.title = "GSTR-1 Sales"
        ws1.merge_cells("A1:H1")
        ws1["A1"] = "UNIVERIN PRIVATE LIMITED — GSTR-1 Sales Register — " + month_name
        ws1["A1"].font = Font(bold=True, size=13, color="2563EB")
        ws1["A1"].alignment = Alignment(horizontal="center")
        headers = ["Order ID","Date","Buyer","Vendor","Taxable Value","CGST","SGST","Total"]
        widths  = [20,15,20,20,15,12,12,15]
        for col,(h,w) in enumerate(zip(headers,widths),1):
            cell = ws1.cell(row=3,column=col,value=h); sh(cell)
            ws1.column_dimensions[get_column_letter(col)].width = w
        t_tax=Decimal("0"); t_tot=Decimal("0")
        for i,order in enumerate(orders):
            row=4+i; ov=Decimal(str(order.subtotal or 0)); gst=Decimal(str(order.gst_on_platform or 0))
            total=ov+gst+Decimal(str(order.platform_fee or 0))+Decimal(str(order.delivery_fee or 0))
            t_tax+=ov; t_tot+=total
            fill=LIGHT_FILL if i%2==0 else PatternFill("solid",fgColor="FFFFFF")
            for col,(val,align) in enumerate(zip([str(order.id)[:12].upper(),order.created_at.strftime("%d %b %Y"),getattr(order.buyer,"phone_number","N/A"),order.vendor.shop_name,f"Rs.{ov:.2f}",f"Rs.{gst/2:.2f}",f"Rs.{gst/2:.2f}",f"Rs.{total:.2f}"],["left","center","left","left","right","right","right","right"]),1):
                cell=ws1.cell(row=row,column=col,value=val); sc(cell,align=align); cell.fill=fill
        tr=4+len(orders)
        for col,val in enumerate(["TOTAL","","","",f"Rs.{t_tax:.2f}","","",f"Rs.{t_tot:.2f}"],1):
            cell=ws1.cell(row=tr,column=col,value=val); cell.font=Font(bold=True,color="FFFFFF"); cell.fill=BLUE_FILL; cell.alignment=Alignment(horizontal="right" if col>4 else "center")

        # Sheet 2 GSTR-8
        ws2=wb.create_sheet("GSTR-8 TCS")
        ws2.merge_cells("A1:H1")
        ws2["A1"]="UNIVERIN PRIVATE LIMITED — GSTR-8 TCS Register — "+month_name
        ws2["A1"].font=Font(bold=True,size=13,color="F97316"); ws2["A1"].alignment=Alignment(horizontal="center")
        h2=["Order ID","Date","Vendor","Vendor GSTIN","Taxable Value","CGST TCS","SGST TCS","Total TCS"]
        w2=[20,15,20,20,15,15,15,12]
        for col,(h,w) in enumerate(zip(h2,w2),1):
            cell=ws2.cell(row=3,column=col,value=h); sh(cell,ORANGE_FILL)
            ws2.column_dimensions[get_column_letter(col)].width=w
        t_tcs=Decimal("0")
        for i,order in enumerate(orders):
            row=4+i; ov=Decimal(str(order.subtotal or 0)); tcs=Decimal(str(order.tcs_amount or 0)); t_tcs+=tcs
            fill=LIGHT_FILL if i%2==0 else PatternFill("solid",fgColor="FFFFFF")
            for col,(val,align) in enumerate(zip([str(order.id)[:12].upper(),order.created_at.strftime("%d %b %Y"),order.vendor.shop_name,getattr(order.vendor,"gstin","N/A") or "N/A",f"Rs.{ov:.2f}",f"Rs.{tcs/2:.2f}",f"Rs.{tcs/2:.2f}",f"Rs.{tcs:.2f}"],["left","center","left","left","right","right","right","right"]),1):
                cell=ws2.cell(row=row,column=col,value=val); sc(cell,align=align); cell.fill=fill

        # Sheet 3 Commission
        ws3=wb.create_sheet("Commission")
        ws3.merge_cells("A1:I1")
        ws3["A1"]="UNIVERIN PRIVATE LIMITED — Commission Register — "+month_name
        ws3["A1"].font=Font(bold=True,size=13,color="16A34A"); ws3["A1"].alignment=Alignment(horizontal="center")
        h3=["Order ID","Date","Vendor","Category","Order Value","Rate","Commission","GST 18%","Total"]
        w3=[20,15,20,15,15,10,14,12,14]
        for col,(h,w) in enumerate(zip(h3,w3),1):
            cell=ws3.cell(row=3,column=col,value=h); sh(cell,GREEN_FILL)
            ws3.column_dimensions[get_column_letter(col)].width=w
        t_comm=Decimal("0")
        for i,order in enumerate(orders):
            row=4+i; ov=Decimal(str(order.subtotal or 0)); comm=Decimal(str(order.commission_amount or 0)); gst=Decimal(str(order.gst_on_platform or 0)); t_comm+=comm
            fill=LIGHT_FILL if i%2==0 else PatternFill("solid",fgColor="FFFFFF")
            for col,(val,align) in enumerate(zip([str(order.id)[:12].upper(),order.created_at.strftime("%d %b %Y"),order.vendor.shop_name,getattr(order.vendor,"category","N/A") or "N/A",f"Rs.{ov:.2f}",f"{order.commission_rate or 0}%",f"Rs.{comm:.2f}",f"Rs.{gst:.2f}",f"Rs.{comm+gst:.2f}"],["left","center","left","left","right","center","right","right","right"]),1):
                cell=ws3.cell(row=row,column=col,value=val); sc(cell,align=align); cell.fill=fill

        # Sheet 4 Settlement
        ws4=wb.create_sheet("Settlement")
        ws4.merge_cells("A1:G1")
        ws4["A1"]="UNIVERIN PRIVATE LIMITED — Settlement Summary — "+month_name
        ws4["A1"].font=Font(bold=True,size=13,color="7C3AED"); ws4["A1"].alignment=Alignment(horizontal="center")
        h4=["Vendor","Category","Total Orders","Gross Value","Commission","TCS","Net Payout"]
        w4=[25,15,14,15,15,12,15]
        for col,(h,w) in enumerate(zip(h4,w4),1):
            cell=ws4.cell(row=3,column=col,value=h); sh(cell,PatternFill("solid",fgColor="7C3AED"))
            ws4.column_dimensions[get_column_letter(col)].width=w
        vendor_summary=orders.values("vendor__shop_name","vendor__category").annotate(total_orders=Count("id"),gross=Sum("subtotal"),commission=Sum("commission_amount"),tcs=Sum("tcs_amount"))
        for i,v in enumerate(vendor_summary):
            row=4+i; gross=Decimal(str(v["gross"] or 0)); comm=Decimal(str(v["commission"] or 0)); tcs=Decimal(str(v["tcs"] or 0)); net=gross-comm-tcs
            fill=LIGHT_FILL if i%2==0 else PatternFill("solid",fgColor="FFFFFF")
            for col,(val,align) in enumerate(zip([v["vendor__shop_name"],v["vendor__category"] or "N/A",str(v["total_orders"]),f"Rs.{gross:.2f}",f"Rs.{comm:.2f}",f"Rs.{tcs:.2f}",f"Rs.{net:.2f}"],["left","left","center","right","right","right","right"]),1):
                cell=ws4.cell(row=row,column=col,value=val); sc(cell,align=align); cell.fill=fill
                if col==7: cell.font=Font(bold=True,size=10,color="16A34A")

        response=HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"]=f'attachment; filename="univerin_gst_{month_name}.xlsx"'.replace(" ","_")
        wb.save(response)
        return response

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        now = datetime.now()
        months = []
        for i in range(6):
            m = now.month - i
            y = now.year
            if m <= 0:
                m += 12
                y -= 1
            months.append({"month": m, "year": y, "label": datetime(y, m, 1).strftime("%B %Y")})
        extra_context["export_months"] = months
        extra_context["current_month"] = now.month
        extra_context["current_year"] = now.year
        return super().changelist_view(request, extra_context=extra_context)

@admin.register(OrderItem)
class OrderItemAdmin(admin.ModelAdmin):
    list_display  = ['order', 'product', 'quantity', 'price']
    search_fields = ['product__name']

@admin.register(Notification)
class NotificationAdmin(admin.ModelAdmin):
    list_display  = ['user', 'title', 'type', 'is_read', 'created_at']
    list_filter   = ['type', 'is_read']
    search_fields = ['user__full_name', 'title']

@admin.register(Review)
class ReviewAdmin(admin.ModelAdmin):
    list_display  = ['buyer', 'vendor', 'rating', 'created_at']
    list_filter   = ['rating']
    search_fields = ['buyer__full_name', 'vendor__shop_name']

@admin.register(Cart)
class CartAdmin(admin.ModelAdmin):
    list_display  = ['buyer', 'product', 'vendor', 'quantity', 'added_at']
    list_filter   = ['vendor']
    search_fields = ['buyer__full_name', 'product__name']

from .models import Coupon
@admin.register(Coupon)
class CouponAdmin(admin.ModelAdmin):
    list_display  = ['code', 'discount_type', 'discount_value', 'min_order', 'used_count', 'max_uses', 'valid_until', 'is_active']
    list_editable = ['is_active']
    search_fields = ['code']
